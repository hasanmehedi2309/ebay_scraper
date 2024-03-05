from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import requests
import os
import csv
import concurrent.futures
import threading

# name value function


def name_value(value):
    name_text = value.text
    name_list = name_text.split("(")
    if "/" in name_list[0]:
        return name_list[0].replace("/","")
    else:
        return name_list[0]

# save xl file's function
def save_file():
    all_dir = os.listdir()
    if "save.xlsx" not in all_dir:
        wb_creation = Workbook()
        ws_creation = wb_creation.active
        ws_creation.append(["sub-category", "page number", "complete", "page link", "first product status"])
        wb_creation.save("save.xlsx")

# page number func
def subcategory_extraction():
    wb = load_workbook("save.xlsx")
    ws = wb.active
    print(ws)
    all_rows = ws.max_row
    print(all_rows)
    for row in range(1, all_rows+1):
        completed_status = ws.cell(row=row, column=3).value
        print(completed_status)
        if completed_status == "no":
            sub_name = ws.cell(row=row, column=1).value
            print(sub_name)
            first_status = ws.cell(row=row, column=5).value
            print(first_status)
    return [sub_name, first_status]


# product data extraction func
def product_data_extraction(page_url):
    sub_extract = subcategory_extraction()
    sub_name = sub_extract[0]
    first_product = sub_extract[1]
    print(first_product)
    # driver.get(page_url)
    r = requests.get(page_url)
    # print(r.status_code)
    product_page_html = r.content
    # entered into the product page
    # product_page_html = driver.page_source
    product_page_soup = BeautifulSoup(product_page_html, "html.parser")
    try:
        p_title = product_page_soup.find("h1", {"id": "itemTitle", "class": "it-ttl"}).text
        print(p_title)
    except:
        p_title = "Not given"

    try:
        price = product_page_soup.find("span", {"id": "prcIsum"}).text
        print(price)
    except:
        price = "Not given"

    try:
        location_div = product_page_soup.find("div", {"id": "itemLocation"})
        # print(location_div)
        location = location_div.find_all("div")[1].span.text
        print(location)
    except:
        location = "Not given"

    try:
        time.sleep(2.5)
        quantity_tag = product_page_soup.find("span", {"id": "qtySubTxt"}).text
        quantity = quantity_tag.lstrip()
        print(quantity)
    except:
        quantity = "Not given"

    # inserting into all items list
    all_items = [p_title, price, location, quantity]

    # specifications of the item(item specifics) processing
    # item_spec = driver.find_elements_by_css_selector("#viTabs_0_is>div>table>tbody>tr")

    time.sleep(2)
    item_specs = product_page_soup.select("#viTabs_0_is>div>table>tr")
    print(item_specs)

    item_info = []  # all the item specifics are stored in here including heading

    for item in item_specs:
        try:
            heading = item.find_all("td")
            # print(heading)
            heading1 = heading[0].text
            heading1 = heading1.lstrip()
            # print(heading1.lstrip())
            value1 = heading[1].find("span").text
            print(heading1 + value1)
            item_info.append(heading1 + value1)
        except:
            pass
        try:
            heading2 = heading[2].text
            heading2 = heading2.lstrip()
            # print(heading1.lstrip())
            value2 = heading[3].find("span").text
            print(value2)
            item_info.append(heading2 + value2)
        except:
            pass
    # print(f"item info{item_info}")

    for item in item_info:
        # item_heading_value = item.split(":")
        all_items.append(item)
    # print(f"item heading {all_items}")
    # all_items_list = all_items_list + all_items
    # print(all_items_list)
    all_items_list.append(all_items)
    return

driver = webdriver.Chrome(ChromeDriverManager().install())

driver.get("http://www.ebaystores.com/Car-Parts-Wholesale/A-C-Condensers-/_i.html?_fsub=5827849&_sid=139020616&_trksid=p4634.c0.m322")

print(driver.title)

first_product = True
program_started = True
all_items = []
all_items_list = []
# categories finding
main_page_soup = BeautifulSoup(driver.page_source, "html.parser")
subcategories = main_page_soup.select("div.lcat ul ul li")

# save file creation
save_file()
for subcategory in subcategories:
    # subcategory name

    subcategory_name = name_value(subcategory)
    print(f"sub name: {subcategory_name}")

    # subcategory link extraction
    sub_partial_link = subcategory.find("a")["href"]
    print(sub_partial_link)
    sub_link = "http://www.ebaystores.com" + sub_partial_link
    print(sub_link)
    # driver.get(sub_link)

    # load file
    if program_started is True:
        print("programme started true")
        l_wb = load_workbook("save.xlsx")
        l_ws = l_wb.active
        save_file_max_row = l_ws.max_row
        all_saved_sub_cat = []
        for i in range(1, save_file_max_row + 1):
            val = l_ws.cell(row=i, column=1).value
            all_saved_sub_cat.append(val)
        if subcategory_name not in all_saved_sub_cat:
            l_ws.append([subcategory_name, 1, "no", None, "true"])
            page_num = l_ws.cell(row=save_file_max_row + 1, column=2).value
            l_wb.save("save.xlsx")
            print(f"{subcategory_name} not found")
        else:
            index = 1
            print(f"{subcategory_name} found")
            for saved_cat in all_saved_sub_cat:
                if saved_cat == subcategory_name:
                    break
                else:
                    index += 1
            if l_ws.cell(row=index, column=3).value == "yes":
                continue
            else:
                page_num = l_ws.cell(row=index, column=2).value
                print(page_num)
                url = l_ws.cell(row=index, column=4).value
                l_wb.save("save.xlsx")

                first_product = False
        program_started = False
    else:
        print("Programme not started")
        l_wb = load_workbook("save.xlsx")
        l_ws = l_wb.active
        l_ws.append([subcategory_name, 1, "no", None, "true"])
        page_num = 1
        l_wb.save("save.xlsx")
        first_product = True

    if page_num == 1:
        driver.get(sub_link)
    else:
        driver.get(url)

    for i in range(1, 500):
        # search result page
        search_page_html = driver.page_source
        soup = BeautifulSoup(search_page_html, "html.parser")
        # print(soup)

        # current page link
        cur_page_link = driver.current_url
        print(f"current link : {cur_page_link}")

        # finding product link
        # div containing all the products in the page

        all_products_div = soup.find("div", {"id": "lvc"})
        # print(all_products_div)
        if all_products_div is not None:
            pass
        else:
            break

        product_tables = all_products_div.find_all("table")
        # print(product_table)

        prod_link_list = []
        for product_table in product_tables:
            # product link to go to the product page
            product_link = product_table.find_all("a")[1]["href"]
            # print(product_link)
            # driver.get(product_link)
            prod_link_list.append(product_link)
        print(len(prod_link_list))

        if first_product is True:
            item_info = []
            driver.get(prod_link_list[0])
            headers_list = ["product title", "price", "location", "quantity"]
            item_spec = driver.find_elements_by_css_selector("#viTabs_0_is>div>table>tbody>tr")
            for spec in item_spec:
                heading = spec.find_elements_by_tag_name("td")
                try:
                    heading1 = heading[0].text
                    value1 = heading[1].find_element_by_tag_name("span").text
                    print(heading1 + value1)
                    item_info.append(heading1 + value1)
                except:
                    pass

                try:
                    heading2 = heading[2].text
                    value2 = heading[3].find_element_by_tag_name("span").text
                    print(heading2 + value2)
                    item_info.append(heading2 + value2)

                except:
                    pass

            for spec_header in item_info:
                spec_header_value = spec_header.split(":")
                headers_list.append(spec_header_value[0])
            with open(f"{subcategory_name}.csv", "w", newline="", encoding="utf-8") as f:
                csv_writer = csv.writer(f)
                csv_writer.writerow(headers_list)
            first_product = False
            driver.back()

        # data extraction
        with concurrent.futures.ThreadPoolExecutor() as executor:
            print("inside executer")
            results = executor.map(product_data_extraction, prod_link_list)

        #             for result in results:
        #                 print(result)

        with open(f"{subcategory_name}.csv", "r") as f:
            reader = csv.DictReader(f)
            item_heading = reader.fieldnames
            item_heading_len = len(item_heading)

        # inserting price,title etc info to the csv file
        for one_prod in all_items_list:
            all_info_list = [one_prod[0], one_prod[1], one_prod[2], one_prod[3]]
            for col in range(4, item_heading_len):
                info_present_status = "no"
                col_heading = item_heading[col]
                for product in one_prod:
                    p_h = product.split(":")
                    if p_h[0] == col_heading:
                        all_info_list.append(p_h[1])
                        info_present_status = "yes"
                        break
                    else:
                        info_present_status = "no"
                if info_present_status == "no":
                    all_info_list.append("Not given")
            # item appending
            with open(f"{subcategory_name}.csv", "a", newline="", encoding="utf-8") as f:
                csv_writer = csv.writer(f)
                csv_writer.writerow(all_info_list)

        # emptinging the so that it wont register the old data again
        all_items = []
        all_items_list = []

        # back to the all product page and searching for the next page
        try:

            next_page = driver.find_element_by_link_text("Next")
            next_page_link = next_page.get_attribute("href")
            if next_page_link is None:
                save_wb = load_workbook("save.xlsx")
                save_ws = save_wb.active
                all_rows = save_ws.max_row
                save_ws.cell(row=all_rows, column=3).value = "yes"
                save_wb.save("save.xlsx")
                break
            else:
                save_wb = load_workbook("save.xlsx")
                save_ws = save_wb.active
                all_rows = save_ws.max_row
                save_ws.cell(row=all_rows, column=2).value = page_num + 1
                save_ws.cell(row=all_rows, column=3).value = "no"
                save_ws.cell(row=all_rows, column=4).value = next_page_link
                save_wb.save("save.xlsx")
                next_page.click()
        except NoSuchElementException:
            save_wb = load_workbook("save.xlsx")
            save_ws = save_wb.active
            all_rows = save_ws.max_row
            save_ws.cell(row=all_rows, column=3).value = "yes"
            save_wb.save("save.xlsx")
            break

driver.quit()