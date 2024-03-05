from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import requests
import os
import csv

file_list = os.listdir()
input_file = None
for file in file_list:
    if ".csv" in file:
        input_file = f"{file}"
        print(input_file)
        break
    else:
        print(file)

if input_file is None:
    print("No file found")
    raise AttributeError

driver = webdriver.Chrome(ChromeDriverManager().install())

driver.get("http://www.ebaystores.com/carpartswholesale")

print(driver.title)

first_product = True
with open(input_file, "r", encoding="utf-8") as csv_file:
    csv_reader = csv.reader(csv_file)
    sku_list = []
    for sku in csv_reader:
        sku_list.append(sku)
for sku in sku_list:
    # search button
    search_bar = driver.find_element_by_id("v4-1")
    search_bar.clear()
    search_bar.send_keys(sku[0])
    search_button = driver.find_element(By.NAME,"submit").click()

    for i in range(1, 200):
        # search result page
        search_page_html = driver.page_source
        soup = BeautifulSoup(search_page_html, "html.parser")
        # print(soup)

        # finding product link
        # div containing all the products in the page

        all_products_div = soup.find("div", {"id": "lvc"})
        # print(all_products_div)
        print(all_products_div)
        if all_products_div is not None:
            pass
        else:
            break

        product_tables = all_products_div.find_all("table")
        # print(product_table)

        for product_table in product_tables:
            # image download
            try:
                product_img = product_table.find_all("a")[0].img["src"]
                r = requests.get(product_img)
                with open("output.jpg", "wb") as file:
                    file.write(r.content)
            except:
                product_img = "Not given"

            # product link to go to the product page
            try:
                product_link = product_table.find_all("a")[1]["href"]
                print(product_link)
                driver.get(product_link)
            except:
                break

            # entered into the product page
            product_page_html = driver.page_source
            product_page_soup = BeautifulSoup(product_page_html, "html.parser")
            try:
                p_title = product_page_soup.find("h1", {"id": "itemTitle", "class": "it-ttl"}).text
                print(p_title)
            except:
                p_title  = "Not given"

            try:
                price = product_page_soup.find("span", {"id": "prcIsum"}).text
                print(price)
            except:
                price = "Not given"

            try:
                location_div = product_page_soup.find("div", {"id": "itemLocation"})
                # print(location_div)
                location = location_div.find_all("div")[1].span.text
                print(location)
            except:
                location = "Not given"

            try:
                quantity = driver.find_element_by_id("qtySubTxt").text
                print(quantity)
            except:
                quantity = "Not given"

            # specifications of the item(item specifics) processing
            item_spec = driver.find_elements_by_css_selector("#viTabs_0_is>div>table>tbody>tr")

            item_info = []  # all the item specifics are stored in here including heading

            for spec in item_spec:
                heading = spec.find_elements_by_tag_name("td")
                try:
                    heading1 = heading[0].text
                    value1 = heading[1].find_element_by_tag_name("span").text
                    print(heading1 + value1)
                    item_info.append(heading1 + value1)
                except:
                    pass

                try:
                    heading2 = heading[2].text
                    value2 = heading[3].find_element_by_tag_name("span").text
                    print(heading2 + value2)
                    item_info.append(heading2 + value2)

                except:
                    pass

            if first_product is True:
                wb = Workbook()
                ws = wb.active
                item_heading = ["SKU","Image", "product title", "price", "location", "quantity"]

                for item in item_info:
                    item_heading_value = item.split(":")
                    item_heading.append(item_heading_value[0])

                ws.append(item_heading)
                wb.save(" car parts wholesale.xlsx")
                first_product = False

            workbook = load_workbook(" car parts wholesale.xlsx")
            work_sheet = workbook.active
            max_row = work_sheet.max_row
            max_col = work_sheet.max_column

            # sku input in the xl file
            work_sheet.cell(row=max_row + 1, column=1).value = sku[0]

            # image inserting to the xl file
            img = Image("output.jpg")
            work_sheet.add_image(img,f"B{max_row+1}")

            # inserting price,title etc info to the xl file
            work_sheet.cell(row=max_row+1, column=3).value = p_title
            work_sheet.cell(row=max_row + 1, column=4).value = price
            work_sheet.cell(row=max_row + 1, column=5).value = location
            work_sheet.cell(row=max_row + 1, column=6).value = quantity

            for col in range(7, max_col + 1):
                col_heading = work_sheet.cell(row=1, column=col).value
                for product in item_info:
                    p_h = product.split(":")
                    if p_h[0] == col_heading:
                        work_sheet.cell(row=max_row + 1, column=col).value = p_h[1]

            workbook.save(" car parts wholesale.xlsx")
            driver.back()
            time.sleep(2)
        # back to the all product page and searching for the next page
        try:
            next_page = driver.find_element_by_link_text("Next")
        except:
            next_page = None
        if next_page is None:
            break
        else:
            next_page.click()

driver.quit()



